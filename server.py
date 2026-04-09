from __future__ import annotations

import cgi
import io
import json
import os
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer

from openpyxl import load_workbook


ROOT = os.path.dirname(os.path.abspath(__file__))
PREFERRED_SHEETS = {"0331", "0228", "3月版本积分排名"}


def normalize_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def find_header_row(rows):
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:10]):
        score = sum(1 for item in row if item)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def worksheet_to_records(worksheet):
    raw_rows = [[normalize_cell(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
    raw_rows = [row for row in raw_rows if any(row)]
    if not raw_rows:
        return {"headers": [], "rows": []}

    header_index = find_header_row(raw_rows)
    headers = raw_rows[header_index]
    clean_headers = []
    for index, header in enumerate(headers):
        if not header and clean_headers:
            break
        clean_headers.append(header or f"字段{index + 1}")
    width = len(clean_headers)

    records = []
    for row in raw_rows[header_index + 1 :]:
        padded = row[:width] + [""] * max(0, width - len(row))
        record = {clean_headers[index]: padded[index] for index in range(width)}
        if any(record.values()):
            records.append(record)

    return {"headers": clean_headers, "rows": records}


class WorkbookHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=ROOT, **kwargs)

    def do_POST(self):
        if self.path != "/api/parse-workbook":
            self.send_error(HTTPStatus.NOT_FOUND, "Unknown endpoint")
            return

        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type", ""),
            },
        )

        upload = form["file"] if "file" in form else None
        if upload is None or not getattr(upload, "file", None):
            self.send_error(HTTPStatus.BAD_REQUEST, "Missing file")
            return

        selected_sheet = form.getfirst("sheet", "")

        try:
            workbook = load_workbook(io.BytesIO(upload.file.read()), data_only=True)
        except Exception as exc:
            self.send_error(HTTPStatus.BAD_REQUEST, f"Invalid workbook: {exc}")
            return

        sheets = workbook.sheetnames
        active_sheet = selected_sheet if selected_sheet in sheets else next(
            (name for name in sheets if name.strip() in PREFERRED_SHEETS),
            sheets[0] if sheets else "",
        )
        payload = worksheet_to_records(workbook[active_sheet]) if active_sheet else {"headers": [], "rows": []}
        body = json.dumps(
            {
                "sheets": sheets,
                "activeSheet": active_sheet,
                "headers": payload["headers"],
                "rows": payload["rows"][:5000],
            },
            ensure_ascii=False,
        ).encode("utf-8")

        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


if __name__ == "__main__":
    server = ThreadingHTTPServer(("127.0.0.1", 8000), WorkbookHandler)
    print("Serving on http://127.0.0.1:8000")
    server.serve_forever()
